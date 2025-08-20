# -*- coding: utf-8 -*-
"""
資料庫操作模組
處理Excel檔案的讀寫操作
"""

import os
from datetime import datetime
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import tempfile

class DatabaseManager:
    def __init__(self, base_path):
        """
        初始化資料庫管理器
        
        Args:
            base_path (str): 基礎路徑
        """
        self.base_path = base_path
        self.database_file = os.path.join(base_path, 'OIR_database.xlsx')
        self.sample_file = os.path.join(base_path, '..', 'OIR_Report_Sample.xlsx')
        
        # 確保資料庫檔案存在
        self._ensure_database_exists()
    
    def _ensure_database_exists(self):
        """確保資料庫檔案存在，如果不存在則創建"""
        if not os.path.exists(self.database_file):
            self._create_database()
    
    def _create_database(self):
        """創建新的資料庫檔案"""
        wb = Workbook()
        
        # 創建OIS工作表
        ws_ois = wb.active
        ws_ois.title = "OIS"
        
        # OIS工作表標題
        ois_headers = [
            'OIS No.', 'Model Code', 'Model Desc.', 'Model Version', 
            'Item', 'SC Symbol', 'Description', 'Minimum Limit', 
            'Maximum Limit', 'Median', 'Unit', 'A.QAL(%) of Sample Size', 
            'Type of Data', 'Measurement Equipment'
        ]
        
        for col, header in enumerate(ois_headers, 1):
            ws_ois.cell(row=1, column=col, value=header)
        
        # 嘗試從OIR_Report_Sample.xlsx匯入Standards數據
        sample_file = os.path.join(self.base_path, '..', 'OIR_Report_Sample.xlsx')
        if os.path.exists(sample_file):
            try:
                self._import_standards_from_sample(ws_ois, sample_file)
            except Exception as e:
                print(f"Warning: Could not import standards from sample file: {e}")
                # 如果匯入失敗，使用預設示例數據
                self._add_default_sample_data(ws_ois)
        else:
            # 如果樣本檔案不存在，使用預設示例數據
            self._add_default_sample_data(ws_ois)
        
        # 創建database工作表
        ws_db = wb.create_sheet("database")
        
        # database工作表標題
        db_headers = [
            'Date', 'Model No.', 'Model Description', 'OIS No.', 'Lot No.', 
            'Item', 'Datapoint_1', 'Datapoint_2', 'Datapoint_3', 'Datapoint_4', 
            'Datapoint_5', 'Datapoint_6', 'Datapoint_7', 'Datapoint_8', 
            'Datapoint_9', 'Datapoint_10', 'Operator'
        ]
        
        for col, header in enumerate(db_headers, 1):
            ws_db.cell(row=1, column=col, value=header)
        
        # 保存檔案
        wb.save(self.database_file)
    
    def _import_standards_from_sample(self, ws_ois, sample_file):
        """從OIR_Report_Sample.xlsx匯入Standards數據"""
        from openpyxl import load_workbook
        
        # 載入樣本檔案
        sample_wb = load_workbook(sample_file)
        if 'Standards' not in sample_wb.sheetnames:
            raise Exception("Standards sheet not found in sample file")
        
        sample_ws = sample_wb['Standards']
        
        # 轉換Standards格式到OIS格式
        # Standards格式: Item, OIS No., Model Desc, OIS Rev., Model Revision, SC Symbol, Description, Minimum Limit, Maximum Limit, Median, Unit, A.QAL(%) of Sample Size, Type of Data, Measurement Equipment
        # OIS格式: OIS No., Model Code, Model Desc., Model Version, Item, SC Symbol, Description, Minimum Limit, Maximum Limit, Median, Unit, A.QAL(%) of Sample Size, Type of Data, Measurement Equipment
        
        row_num = 2  # 從第2行開始寫入數據
        
        for row in sample_ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 跳過空行
                continue
                
            # 轉換數據格式
            ois_row = [
                row[1],   # OIS No.
                row[2],   # Model Code (Model Desc)
                row[2],   # Model Desc.
                row[4],   # Model Version (Model Revision)
                row[0],   # Item
                row[5],   # SC Symbol
                row[6],   # Description
                row[7],   # Minimum Limit
                row[8],   # Maximum Limit
                row[9],   # Median
                row[10],  # Unit
                row[11],  # A.QAL(%) of Sample Size
                row[12],  # Type of Data
                row[13]   # Measurement Equipment
            ]
            
            # 寫入OIS工作表
            for col, value in enumerate(ois_row, 1):
                ws_ois.cell(row=row_num, column=col, value=value)
            
            row_num += 1
    
    def _add_default_sample_data(self, ws_ois):
        """添加預設示例數據"""
        sample_data = [
            ['DCCDC-IS-11301110', '1999-1130111', 'DCJ72(4)MLG-1130111', 'E', 1, '', 'dimension_1', 10.0, 15.0, 12.5, 'mm', 10, 'A', 'Caliper'],
            ['DCCDC-IS-11301110', '1999-1130111', 'DCJ72(4)MLG-1130111', 'E', 2, '', 'dimension_2', 5.0, 8.0, 6.5, 'mm', 10, 'A', 'Caliper'],
            ['DCCDC-IS-11301110', '1999-1130111', 'DCJ72(4)MLG-1130111', 'E', 3, '', 'dimension_3', 2.0, 4.0, 3.0, 'mm', 10, 'A', 'Height indicator'],
            ['DCCDC-IS-11301110', '1999-1130111', 'DCJ72(4)MLG-1130111', 'E', 4, '', 'dimension_4', 1.0, 2.5, 1.75, 'mm', 10, 'A', 'Projector'],
            ['DCCDC-IS-11301110', '1999-1130111', 'DCJ72(4)MLG-1130111', 'E', 5, '', 'dimension_5', 0.5, 1.0, 0.75, 'mm', 10, 'A', 'Projector']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws_ois.cell(row=row, column=col, value=value)
    
    def get_ois_data(self, ois_no):
        """
        根據OIS編號獲取OIS數據
        
        Args:
            ois_no (str): OIS編號
            
        Returns:
            list: OIS數據列表，如果找不到則返回空列表
        """
        try:
            wb = load_workbook(self.database_file)
            ws = wb['OIS']
            
            # 獲取標題行
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 查找匹配的OIS編號
            ois_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == ois_no:  # OIS No. 在第一欄
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i]
                        else:
                            row_dict[header] = None
                    ois_data.append(row_dict)
            
            wb.close()
            return ois_data
            
        except Exception as e:
            print(f"Error reading OIS data: {e}")
            return []
    
    def get_model_description(self, model_code):
        """
        根據型號代碼獲取型號描述
        
        Args:
            model_code (str): 型號代碼
            
        Returns:
            str: 型號描述
        """
        try:
            wb = load_workbook(self.database_file)
            ws = wb['OIS']
            
            # 查找Model Code欄位的索引
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            model_code_col = headers.index('Model Code') if 'Model Code' in headers else 1
            model_desc_col = headers.index('Model Desc.') if 'Model Desc.' in headers else 2
            
            # 查找匹配的型號代碼
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > model_code_col and row[model_code_col] == model_code:
                    if len(row) > model_desc_col and row[model_desc_col]:
                        wb.close()
                        return row[model_desc_col]
            
            wb.close()
            return model_code
            
        except Exception as e:
            print(f"Error getting model description: {e}")
            return model_code
    
    def save_inspection_data(self, data):
        """
        保存檢驗數據到資料庫
        
        Args:
            data (dict): 檢驗數據
                - date: 日期
                - model_no: 型號
                - model_desc: 型號描述
                - ois_no: OIS編號
                - lot_no: 批號
                - items: 項目數據列表
                - operator: 操作員
        
        Returns:
            bool: 保存成功返回True，失敗返回False
        """
        try:
            wb = load_workbook(self.database_file)
            ws = wb['database']
            
            # 獲取下一行的行號
            next_row = ws.max_row + 1
            
            # 為每個項目添加記錄
            for item_data in data['items']:
                # 基本資料
                ws.cell(row=next_row, column=1, value=data['date'])          # Date
                ws.cell(row=next_row, column=2, value=data['model_no'])      # Model No.
                ws.cell(row=next_row, column=3, value=data['model_desc'])    # Model Description
                ws.cell(row=next_row, column=4, value=data['ois_no'])        # OIS No.
                ws.cell(row=next_row, column=5, value=data['lot_no'])        # Lot No.
                ws.cell(row=next_row, column=6, value=item_data['item'])     # Item
                
                # 添加10個數據點 (columns 7-16)
                datapoints = item_data.get('datapoints', [])
                for i in range(10):
                    col = 7 + i  # Datapoint_1 to Datapoint_10
                    value = datapoints[i] if i < len(datapoints) else None
                    ws.cell(row=next_row, column=col, value=value)
                
                ws.cell(row=next_row, column=17, value=data['operator'])     # Operator
                
                next_row += 1
            
            # 保存檔案
            wb.save(self.database_file)
            wb.close()
            
            return True
            
        except Exception as e:
            print(f"Error saving inspection data: {e}")
            return False
    
    def search_history_data(self, filters):
        """
        搜尋歷史數據
        
        Args:
            filters (dict): 搜尋條件
                - date_from: 起始日期
                - date_to: 結束日期
                - model_no: 型號
                - lot_no: 批號
                - operator: 操作員
        
        Returns:
            list: 搜尋結果列表
        """
        try:
            wb = load_workbook(self.database_file)
            ws = wb['database']
            
            # 獲取標題行
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 獲取所有數據
            results = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:  # 跳過空行
                    continue
                
                # 建立記錄字典
                record = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        value = row[i]
                        # 特別處理日期欄位，確保格式一致
                        if header == 'Date' and value:
                            if isinstance(value, datetime):
                                record[header] = value.strftime('%Y-%m-%d')
                            else:
                                record[header] = str(value)
                        else:
                            record[header] = value
                    else:
                        record[header] = None
                
                # 應用篩選條件
                match = True
                
                # 日期篩選
                if filters.get('date_from') or filters.get('date_to'):
                    try:
                        record_date = datetime.strptime(str(record['Date']), '%Y-%m-%d') if record.get('Date') else None
                        if record_date:
                            if filters.get('date_from'):
                                filter_date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d')
                                if record_date < filter_date_from:
                                    match = False
                            if filters.get('date_to'):
                                filter_date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d')
                                if record_date > filter_date_to:
                                    match = False
                    except:
                        pass
                
                # 型號篩選
                if filters.get('model_no') and match:
                    model_no = str(record.get('Model No.', ''))
                    if filters['model_no'].lower() not in model_no.lower():
                        match = False
                
                # 批號篩選
                if filters.get('lot_no') and match:
                    lot_no = str(record.get('Lot No.', ''))
                    if filters['lot_no'].lower() not in lot_no.lower():
                        match = False
                
                # 操作員篩選
                if filters.get('operator') and match:
                    if record.get('Operator') != filters['operator']:
                        match = False
                
                if match:
                    results.append(record)
            
            wb.close()
            return results
            
        except Exception as e:
            print(f"Error searching history data: {e}")
            return []
    
    def create_report_excel(self, data, template_file=None):
        """
        創建報告Excel檔案，基於OIR_Report_Sample.xlsx的確切結構
        
        Args:
            data (dict): 報告數據
            template_file (str): 模板檔案路徑，如果為None則使用預設模板
        
        Returns:
            str: 創建的臨時檔案路徑，如果失敗則返回None
        """
        try:
            # 使用模板檔案
            if template_file and os.path.exists(template_file):
                # 複製模板檔案到臨時位置
                temp_dir = tempfile.gettempdir()
                report_filename = f"OIR_{datetime.now().strftime('%Y%m%d')}_{data.get('model_no', 'UNKNOWN')}_{data.get('ois_no', 'UNKNOWN')}.xlsx"
                temp_file = os.path.join(temp_dir, report_filename)
                
                print(f"Creating report: {report_filename}")
                print(f"Template file: {template_file}")
                print(f"Temp file: {temp_file}")
                
                # 確保目標目錄存在
                os.makedirs(os.path.dirname(temp_file), exist_ok=True)
                shutil.copy2(template_file, temp_file)
                
                # 載入並填充數據到Dimension_Inspection_Report工作表
                wb = load_workbook(temp_file)
                
                # 選擇Sheet1工作表（新模板的工作表名稱）
                if 'Sheet1' in wb.sheetnames:
                    ws = wb['Sheet1']
                else:
                    ws = wb.active
                
                # 確保只有一個工作表（新模板本來就只有一個）
                print(f"Available sheets: {wb.sheetnames}")
                print(f"Using sheet: {ws.title}")
                
                # 填充基本資訊 - 根據確切的儲存格位置
                ws['D3'] = data.get('model_no', '')          # Model No.
                ws['D4'] = data.get('order_no', '')          # Order No.
                ws['D5'] = data.get('shipment_size', '')     # Shipment Size
                ws['D6'] = data.get('lot_no', '')            # Lot No.
                ws['K3'] = datetime.now().strftime('%Y-%m-%d')  # Today's Date
                ws['K4'] = data.get('inspector', '')         # Inspected By
                ws['K5'] = data.get('location', '')          # Location
                ws['K6'] = data.get('ois_no', '')            # OIS No.
                
                # 日期使用Excel公式 =TODAY() (已經在模板中)
                # ws['K3'] 應該已經有 =TODAY() 公式
                
                # 填充檢驗數據 - 從A9開始，數據從C9到L9
                items_data = data.get('items', [])
                for i, item_data in enumerate(items_data):
                    row = 9 + i  # 從第9行開始
                    
                    # Item No. (A欄)
                    ws[f'A{row}'] = item_data.get('item', i + 1)
                    
                    # Standard欄位(B欄) - 留空
                    ws[f'B{row}'] = ''
                    
                    # 填充10個測量數值 (C9到L9)
                    datapoints = item_data.get('datapoints', [])
                    for j, datapoint in enumerate(datapoints):
                        if j < 10:  # 最多10個數據點
                            col_letter = chr(ord('C') + j)  # C, D, E, F, G, H, I, J, K, L
                            if datapoint is not None:
                                ws[f'{col_letter}{row}'] = datapoint
                    
                    # Result欄位(M, N欄) - 留空
                    ws[f'M{row}'] = ''  # Accept
                    ws[f'N{row}'] = ''  # Reject
                
                wb.save(temp_file)
                return temp_file
                
            else:
                # 如果沒有模板檔案，創建基本報告
                temp_dir = tempfile.gettempdir()
                report_filename = f"OIR_{datetime.now().strftime('%Y%m%d')}_{data['model_no']}_{data['ois_no']}.xlsx"
                temp_file = os.path.join(temp_dir, report_filename)
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Dimension_Inspection_Report"
                
                # 創建基本的報告結構（模仿OIR_Report_Sample.xlsx）
                ws['A1'] = 'Dimension Inspection Report'
                ws['A1'].font = Font(size=16, bold=True)
                ws['A2'] = '尺寸檢驗報告'
                
                # 基本資訊標籤
                ws['A3'] = 'Model No.產品型號:'
                ws['A4'] = 'Order No. 訂單編號:'
                ws['A5'] = 'Shipment Size 出貨數量:'
                ws['A6'] = 'Lot No. 批號:'
                ws['I3'] = 'Date 日期:'
                ws['I4'] = 'Inspected By 測量人:'
                ws['I5'] = 'Location 位置:'
                ws['I6'] = 'OIS No. OIS 編號:'
                
                # 填充數據
                ws['C3'] = data.get('model_no', '')
                ws['C4'] = data.get('order_no', '')
                ws['C5'] = data.get('shipment_size', '')
                ws['C6'] = data.get('lot_no', '')
                ws['K3'] = '=TODAY()'
                ws['K4'] = data.get('inspector', '')
                ws['K5'] = data.get('location', '')
                ws['K6'] = data.get('ois_no', '')
                
                # 表格標題
                ws['A7'] = 'Item No.\n序號'
                ws['B7'] = 'Standard\n標準'
                ws['C7'] = 'Readings of 10 Measurements 10個測量數值'
                ws['M7'] = 'Result 結果'
                
                # 數字標題 1-10
                for i in range(10):
                    col_letter = chr(ord('C') + i)
                    ws[f'{col_letter}8'] = i + 1
                
                ws['M8'] = 'Accept\n接受'
                ws['N8'] = 'Reject\n拒收'
                
                # 清除數據驗證區域的所有驗證規則（C9:L18範圍）
                for row in range(9, 19):  # 假設最多10個項目
                    for col in range(ord('C'), ord('M')):  # C到L列
                        cell = ws[chr(col) + str(row)]
                        if cell.data_validation:
                            cell.data_validation = None
                
                # 填充檢驗數據 - 填充Item No.和實際數據點
                items_data = data.get('items', [])
                for i, item_data in enumerate(items_data):
                    row = 9 + i
                    ws[f'A{row}'] = item_data.get('item', i + 1)
                    
                    # 填入實際數據點到C-L列，保持模板格式
                    datapoints = item_data.get('datapoints', [])
                    for j, datapoint in enumerate(datapoints):
                        if j < 10:  # 只填入前10個數據點
                            col_letter = chr(ord('C') + j)  # C, D, E, ..., L
                            if datapoint is not None:
                                cell = ws[f'{col_letter}{row}']
                                cell.value = float(datapoint) if isinstance(datapoint, (int, float, str)) and str(datapoint).replace('.', '').replace('-', '').isdigit() else datapoint
                                # 清除數據驗證規則以避免紅色顯示
                                cell.data_validation = None
                
                wb.save(temp_file)
                
                # 設置臨時文件5分鐘後自動清理
                import threading
                def cleanup_after_delay():
                    import time
                    time.sleep(300)  # 5分鐘 = 300秒
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                            print(f"Temporary file cleaned up: {temp_file}")
                    except Exception as e:
                        print(f"Error cleaning up temp file: {e}")
                
                cleanup_thread = threading.Thread(target=cleanup_after_delay)
                cleanup_thread.daemon = True
                cleanup_thread.start()
                
                return temp_file
            
        except Exception as e:
            print(f"Error creating report Excel: {e}")
            return None
    
    def cleanup_temp_files(self, file_path):
        """
        清理臨時檔案
        
        Args:
            file_path (str): 要清理的檔案路徑
        """
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Error cleaning up temp file: {e}")

