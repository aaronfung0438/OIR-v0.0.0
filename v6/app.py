# -*- coding: utf-8 -*-
"""
OIR報告系統主應用程式
Flask Web應用，支援多語言介面
"""

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import os
from datetime import datetime
import tempfile
import json
from database import DatabaseManager
from languages import get_text, get_available_languages, get_language_name
from temp_data import TempDataManager
import logging

# 設置日誌
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

# 設置絕對路徑
BASE_PATH = r'C:\Users\aaron\OneDrive\桌面\intern\JE\OIR Report\v6'
app.config['BASE_PATH'] = BASE_PATH

# 初始化資料庫和臨時數據管理器
db_manager = DatabaseManager(BASE_PATH)
temp_manager = TempDataManager(BASE_PATH)

@app.before_request
def before_request():
    """每個請求前的處理"""
    # 設置預設語言
    if 'language' not in session:
        session['language'] = 'en'

@app.context_processor
def inject_template_vars():
    """注入模板變數"""
    return {
        'get_text': lambda key: get_text(key, session.get('language', 'en')),
        'current_lang': session.get('language', 'en'),
        'available_languages': get_available_languages(),
        'get_language_name': lambda code: get_language_name(code, session.get('language', 'en'))
    }

@app.route('/')
def index():
    """首頁"""
    return render_template('index.html')

@app.route('/set_language/<language>')
def set_language(language):
    """設置語言"""
    if language in get_available_languages():
        session['language'] = language
    return redirect(request.referrer or url_for('index'))

@app.route('/new_report')
def new_report():
    """新增報告 - 第一步：輸入基本資訊"""
    return render_template('new_report_step1.html')

@app.route('/new_report/step1', methods=['POST'])
def new_report_step1():
    """處理第一步表單提交"""
    model_no = request.form.get('model_no', '').strip()
    ois_no = request.form.get('ois_no', '').strip()
    inspector = request.form.get('inspector', '').strip()
    
    if not all([model_no, ois_no, inspector]):
        flash(get_text('invalid_input', session.get('language', 'en')), 'error')
        return redirect(url_for('new_report'))
    
    # 檢查OIS編號是否存在 - 必須存在才能繼續
    ois_data = db_manager.get_ois_data(ois_no)
    if not ois_data:
        flash(get_text('ois_not_found', session.get('language', 'en')), 'error')
        return redirect(url_for('new_report'))
    
    # 確保OIS數據是可序列化的
    serializable_ois_data = []
    for item in ois_data:
        serializable_item = {}
        for key, value in item.items():
            # 轉換所有值為可序列化的基本類型
            if value is None:
                serializable_item[key] = None
            elif isinstance(value, (int, float, str, bool)):
                serializable_item[key] = value
            else:
                serializable_item[key] = str(value)
        serializable_ois_data.append(serializable_item)
    
    # 儲存到session
    session['report_data'] = {
        'model_no': model_no,
        'ois_no': ois_no,
        'inspector': inspector,
        'ois_items': serializable_ois_data,
        'current_item': 0,
        'items_data': {}
    }
    
    return redirect(url_for('data_input'))

@app.route('/data_input')
def data_input():
    """數據輸入頁面"""
    if 'report_data' not in session:
        # 嘗試從臨時文件恢復數據
        session_id = session.get('session_id', request.remote_addr)
        temp_data = temp_manager.load_session_data(session_id)
        if temp_data:
            session['report_data'] = temp_data
            flash(get_text('success', session.get('language', 'en')) + ' - Data recovered from temporary storage', 'success')
        else:
            flash(get_text('error', session.get('language', 'en')), 'error')
            return redirect(url_for('new_report'))
    
    report_data = session['report_data']
    current_item = report_data['current_item']
    ois_items = report_data['ois_items']
    
    if current_item >= len(ois_items):
        return redirect(url_for('confirm_data'))
    
    current_item_data = ois_items[current_item]
    
    return render_template('data_input.html', 
                         item_data=current_item_data,
                         current_item=current_item + 1,
                         total_items=len(ois_items),
                         existing_data=report_data['items_data'].get(str(current_item), {}))

@app.route('/data_input/submit', methods=['POST'])
def submit_data_input():
    """提交數據輸入"""
    try:
        logger.info("=== Data input submission started ===")
        logger.info(f"Session keys: {list(session.keys())}")
        
        if 'report_data' not in session:
            logger.error("No report_data in session")
            return jsonify({'success': False, 'message': get_text('error', session.get('language', 'en'))})
        
        report_data = session['report_data']
        current_item = report_data['current_item']
        logger.info(f"Current item: {current_item}")
        logger.info(f"Report data keys: {list(report_data.keys())}")
        
        # 獲取10個數據點
        datapoints = []
        for i in range(1, 11):
            value = request.form.get(f'datapoint_{i}', '').strip()
            logger.info(f"Datapoint {i}: '{value}'")
            if value:
                try:
                    datapoints.append(float(value))
                except ValueError:
                    datapoints.append(value)  # 如果不是數字，保存原始值
            else:
                datapoints.append(None)
        
        logger.info(f"All datapoints: {datapoints}")
    
        # 儲存當前項目的數據 - 確保所有值都是可序列化的
        current_ois_item = report_data['ois_items'][current_item]
        logger.info(f"Current OIS item: {current_ois_item}")
        
        item_data = {
            'item': int(current_ois_item['Item']) if current_ois_item['Item'] is not None else current_item + 1,
            'description': str(current_ois_item['Description']) if current_ois_item['Description'] is not None else '',
            'min_limit': float(current_ois_item['Minimum Limit']) if current_ois_item['Minimum Limit'] is not None else 0.0,
            'max_limit': float(current_ois_item['Maximum Limit']) if current_ois_item['Maximum Limit'] is not None else 0.0,
            'unit': str(current_ois_item['Unit']) if current_ois_item['Unit'] is not None else '',
            'datapoints': datapoints
        }
        
        logger.info(f"Prepared item data: {item_data}")
        # 確保 current_item 是字符串作為字典鍵，避免序列化問題
        report_data['items_data'][str(current_item)] = item_data
        
        # 移動到下一項目
        report_data['current_item'] += 1
        logger.info(f"Moving to next item: {report_data['current_item']}")
        
        # 嘗試保存到session
        logger.info("Attempting to save to session...")
        session['report_data'] = report_data
        logger.info("Session save successful")
        
        # 保存到臨時文件作為備份
        session_id = session.get('session_id', request.remote_addr)
        temp_result = temp_manager.save_session_data(session_id, report_data)
        logger.info(f"Temp file save result: {temp_result}")
        
        # 檢查是否完成所有項目
        if report_data['current_item'] >= len(report_data['ois_items']):
            logger.info("All items completed, redirecting to confirm_data")
            return jsonify({'success': True, 'redirect': url_for('confirm_data')})
        else:
            logger.info("More items remaining, redirecting to data_input")
            return jsonify({'success': True, 'redirect': url_for('data_input')})
            
    except Exception as e:
        logger.error(f"Error in submit_data_input: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/data_input/back')
def data_input_back():
    """返回上一個項目"""
    if 'report_data' not in session:
        return redirect(url_for('new_report'))
    
    report_data = session['report_data']
    if report_data['current_item'] > 0:
        report_data['current_item'] -= 1
        session['report_data'] = report_data
    
    return redirect(url_for('data_input'))

@app.route('/confirm_data')
def confirm_data():
    """確認數據頁面"""
    if 'report_data' not in session:
        flash(get_text('error', session.get('language', 'en')), 'error')
        return redirect(url_for('new_report'))
    
    report_data = session['report_data']
    return render_template('confirm_data.html', report_data=report_data)

@app.route('/confirm_data/submit', methods=['POST'])
def submit_confirm_data():
    """提交確認數據"""
    if 'report_data' not in session:
        return jsonify({'success': False, 'message': get_text('error', session.get('language', 'en'))})
    
    order_no = request.form.get('order_no', '').strip()
    shipment_size = request.form.get('shipment_size', '').strip()
    lot_no = request.form.get('lot_no', '').strip()
    location = request.form.get('location', '').strip()
    
    if not all([order_no, shipment_size, lot_no, location]):
        return jsonify({'success': False, 'message': get_text('invalid_input', session.get('language', 'en'))})
    
    # 更新報告數據
    report_data = session['report_data']
    report_data.update({
        'order_no': order_no,
        'shipment_size': shipment_size,
        'lot_no': lot_no,
        'location': location,
        'date': datetime.now().strftime('%Y-%m-%d')
    })
    session['report_data'] = report_data
    
    return jsonify({'success': True, 'redirect': url_for('preview_report')})

@app.route('/preview_report')
def preview_report():
    """預覽報告"""
    if 'report_data' not in session:
        flash(get_text('error', session.get('language', 'en')), 'error')
        return redirect(url_for('new_report'))
    
    report_data = session['report_data']
    return render_template('preview_report.html', report_data=report_data)

@app.route('/generate_report')
def generate_report():
    """生成並下載Excel報告"""
    if 'report_data' not in session:
        flash(get_text('error', session.get('language', 'en')), 'error')
        return redirect(url_for('new_report'))
    
    report_data = session['report_data']
    
    # 準備數據用於保存到資料庫
    db_data = {
        'date': report_data['date'],
        'model_no': report_data['model_no'],
        'model_desc': db_manager.get_model_description(report_data['model_no']),
        'ois_no': report_data['ois_no'],
        'lot_no': report_data['lot_no'],
        'operator': report_data['inspector'],
        'items': list(report_data['items_data'].values())
    }
    
    try:
        # 保存到資料庫
        if db_manager.save_inspection_data(db_data):
            # 創建Excel報告 - 準備正確的數據格式
            excel_data = {
                'model_no': report_data['model_no'],
                'order_no': report_data.get('order_no', ''),
                'shipment_size': report_data.get('shipment_size', ''),
                'lot_no': report_data.get('lot_no', ''),
                'inspector': report_data['inspector'],
                'location': report_data.get('location', ''),
                'ois_no': report_data['ois_no'],
                'date': report_data.get('date', ''),
                'items': list(report_data['items_data'].values())
            }
            
            # 檢查模板文件是否存在
            sample_file = os.path.join(BASE_PATH, 'OIR_Report_Sample_v2.xlsx')
            logger.info(f"Template file path: {sample_file}")
            logger.info(f"Template file exists: {os.path.exists(sample_file)}")
            
            if not os.path.exists(sample_file):
                logger.error(f"Template file not found: {sample_file}")
                flash('模板文件不存在', 'error')
                return redirect(url_for('preview_report'))
            
            excel_file = db_manager.create_report_excel(excel_data, sample_file)
            logger.info(f"Generated Excel file: {excel_file}")
            
            if excel_file and os.path.exists(excel_file):
                # 清理session
                session.pop('report_data', None)
                flash(get_text('report_generated', session.get('language', 'en')), 'success')
                
                return send_file(excel_file, 
                               as_attachment=True, 
                               download_name=os.path.basename(excel_file),
                               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            else:
                logger.error(f"Failed to create Excel file or file does not exist: {excel_file}")
                flash('Excel文件生成失败', 'error')
                return redirect(url_for('preview_report'))
        else:
            logger.error("Failed to save inspection data to database")
            flash('數據保存失败', 'error')
            return redirect(url_for('preview_report'))
            
    except Exception as e:
        logger.error(f"Error in generate_report: {str(e)}")
        flash(f'報告生成錯誤: {str(e)}', 'error')
        return redirect(url_for('preview_report'))
    
    flash(get_text('file_error', session.get('language', 'en')), 'error')
    return redirect(url_for('preview_report'))



@app.route('/history_report')
def history_report():
    """歷史報告查詢頁面"""
    return render_template('history_report.html')

@app.route('/history_report/search', methods=['POST'])
def search_history():
    """搜尋歷史報告"""
    filters = {
        'date_from': request.form.get('date_from'),
        'date_to': request.form.get('date_to'),
        'model_no': request.form.get('model_no'),
        'lot_no': request.form.get('lot_no'),
        'operator': request.form.get('operator')
    }
    
    # 移除空值
    filters = {k: v for k, v in filters.items() if v}
    
    results = db_manager.search_history_data(filters)
    
    return render_template('history_results.html', results=results, filters=filters)

@app.route('/history_report/prepare', methods=['POST'])
def prepare_history_report():
    """準備歷史報告生成 - 第一步：選擇記錄"""
    try:
        # 從POST請求中獲取記錄數據
        request_data = request.json
        
        if not request_data:
            return jsonify({'success': False, 'message': '無效的記錄數據'})
        
        # 將選中的記錄保存到session中
        session['history_report_data'] = {
            'selected_records': request_data.get('records', []),
            'multiple': request_data.get('multiple', False)
        }
        
        # 返回成功，前端將跳轉到額外資訊輸入頁面
        return jsonify({
            'success': True, 
            'redirect_url': url_for('history_report_additional_info')
        })
        
    except Exception as e:
        logger.error(f"Error preparing history report: {str(e)}")
        return jsonify({'success': False, 'message': f'準備報告錯誤: {str(e)}'})

@app.route('/history_report/additional_info')
def history_report_additional_info():
    """歷史報告額外資訊輸入頁面"""
    if 'history_report_data' not in session:
        flash('無效的會話數據', 'error')
        return redirect(url_for('history_report'))
    
    history_data = session['history_report_data']
    selected_records = history_data['selected_records']
    
    # 從第一筆記錄中獲取基本資訊作為預設值
    if selected_records:
        first_record = selected_records[0]
        default_data = {
            'model_no': first_record.get('Model No.', ''),
            'ois_no': first_record.get('OIS No.', ''),
            'inspector': first_record.get('Operator', ''),
            'lot_no': first_record.get('Lot No.', ''),
        }
    else:
        default_data = {}
    
    return render_template('history_additional_info.html', 
                         selected_count=len(selected_records),
                         default_data=default_data)

@app.route('/history_report/generate', methods=['POST'])
def generate_history_report():
    """根據歷史記錄和額外資訊生成報告"""
    try:
        if 'history_report_data' not in session:
            return jsonify({'success': False, 'message': '無效的會話數據'})
        
        # 獲取額外資訊
        order_no = request.form.get('order_no', '').strip()
        shipment_size = request.form.get('shipment_size', '').strip()
        location = request.form.get('location', '').strip()
        
        history_data = session['history_report_data']
        selected_records = history_data['selected_records']
        
        if not selected_records:
            return jsonify({'success': False, 'message': '沒有選中的記錄'})
        
        # 使用第一筆記錄的基本資訊
        first_record = selected_records[0]
        
        # 準備Excel報告數據格式
        excel_data = {
            'model_no': first_record.get('Model No.', ''),
            'order_no': order_no,
            'shipment_size': shipment_size,
            'lot_no': first_record.get('Lot No.', ''),
            'inspector': first_record.get('Operator', ''),
            'location': location,
            'ois_no': first_record.get('OIS No.', ''),
            'date': first_record.get('Date', ''),
            'items': []
        }
        
        # 重建items數據結構，處理所有選中的記錄
        items_data = []
        for record_index, record in enumerate(selected_records):
            datapoints = []
            for j in range(1, 11):  # 10個數據點
                key = f'Datapoint_{j}'
                if key in record and record[key] is not None:
                    datapoints.append(record[key])
                else:
                    datapoints.append(None)
            
            # 如果有任何非空數據點，添加這個項目
            if any(dp is not None for dp in datapoints):
                items_data.append({
                    'item': record.get('Item', record_index + 1),
                    'description': f'Item {record.get("Item", record_index + 1)}',
                    'min_limit': 0.0,
                    'max_limit': 100.0,
                    'unit': '',
                    'datapoints': datapoints
                })
        
        excel_data['items'] = items_data
        
        logger.info(f"Excel data structure: {list(excel_data.keys())}")
        logger.info(f"Items data type: {type(excel_data['items'])}")
        logger.info(f"Items count: {len(excel_data['items']) if excel_data['items'] else 0}")
        
        # 計算項目數量
        items_count = len(items_data) if items_data else 0
        
        # 檢查模板文件
        sample_file = os.path.join(BASE_PATH, 'OIR_Report_Sample_v2.xlsx')
        if not os.path.exists(sample_file):
            return jsonify({'success': False, 'message': '模板文件不存在'})
        
        # 生成Excel報告
        excel_file = db_manager.create_report_excel(excel_data, sample_file)
        
        if excel_file and os.path.exists(excel_file):
            # 將生成的文件信息保存到session中
            session['generated_report'] = {
                'file_path': excel_file,
                'filename': os.path.basename(excel_file),
                'report_data': excel_data,
                'items_count': items_count
            }
            flash('報告生成成功', 'success')
            
            # 跳轉到報告生成頁面
            return redirect(url_for('history_report_generated'))
        else:
            flash('Excel文件生成失敗', 'error')
            return redirect(url_for('history_report'))
            
    except Exception as e:
        logger.error(f"Error generating history report: {str(e)}")
        flash(f'報告生成錯誤: {str(e)}', 'error')
        return redirect(url_for('history_report'))

@app.route('/history_report/generated')
def history_report_generated():
    """歷史報告生成完成頁面"""
    if 'generated_report' not in session:
        flash('無效的會話數據', 'error')
        return redirect(url_for('history_report'))
    
    report_info = session['generated_report']
    
    return render_template('history_report_generated.html', 
                         report_info=report_info)

@app.route('/download_generated_report')
def download_generated_report():
    """下載生成的報告"""
    if 'generated_report' not in session:
        flash('文件不存在或已過期', 'error')
        return redirect(url_for('history_report'))
    
    report_info = session['generated_report']
    file_path = report_info['file_path']
    filename = report_info['filename']
    
    if os.path.exists(file_path):
        # 清理session
        session.pop('generated_report', None)
        session.pop('history_report_data', None)
        
        return send_file(file_path, 
                        as_attachment=True, 
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        flash('文件不存在或已過期', 'error')
        return redirect(url_for('history_report'))

@app.route('/download_history_report/<filename>')
def download_history_report(filename):
    """下載歷史報告文件"""
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)
    
    if os.path.exists(file_path):
        return send_file(file_path, 
                        as_attachment=True, 
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        flash('文件不存在或已過期', 'error')
        return redirect(url_for('history_report'))

@app.route('/api/ois_items/<ois_no>')
def api_get_ois_items(ois_no):
    """API: 獲取OIS項目"""
    items = db_manager.get_ois_data(ois_no)
    return jsonify(items)

@app.route('/debug')
def debug_info():
    """調試資訊頁面"""
    return render_template('debug_info.html')

@app.route('/api/debug/session')
def api_debug_session():
    """API: 獲取session資訊"""
    return jsonify(dict(session))

@app.route('/api/debug/ois_numbers')
def api_debug_ois_numbers():
    """API: 獲取所有可用的OIS編號"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(db_manager.database_file)
        ws = wb['OIS']
        
        ois_numbers = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # OIS No. is in first column
                ois_numbers.add(row[0])
        
        wb.close()
        return jsonify(sorted(list(ois_numbers)))
        
    except Exception as e:
        logger.error(f"Error getting OIS numbers: {e}")
        return jsonify({'error': str(e)}), 500

@app.errorhandler(404)
def not_found_error(error):
    """404錯誤處理"""
    return render_template('error.html', error_code=404), 404

@app.errorhandler(500)
def internal_error(error):
    """500錯誤處理"""
    return render_template('error.html', error_code=500), 500

if __name__ == '__main__':
    # 確保templates資料夾存在
    templates_dir = os.path.join(BASE_PATH, 'templates')
    if not os.path.exists(templates_dir):
        os.makedirs(templates_dir)
    
    # 確保static資料夾存在
    static_dir = os.path.join(BASE_PATH, 'static')
    if not os.path.exists(static_dir):
        os.makedirs(static_dir)
    
    app.run(debug=True, host='127.0.0.1', port=5000)
